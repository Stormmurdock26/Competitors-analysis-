from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook

from storm_app.workflow import (
    BuildRequest,
    add_brand_rule,
    add_category_rule,
    discover_links,
    normalize_links,
    run_build,
)


def parse_brand_rule(value: str) -> tuple[str, str]:
    if "=" not in value:
        raise argparse.ArgumentTypeError("Brand rules must use Brand=owned or Brand=competitor.")
    brand, ownership = value.split("=", 1)
    ownership = ownership.strip().casefold()
    if ownership not in {"owned", "competitor"}:
        raise argparse.ArgumentTypeError("Brand ownership must be owned or competitor.")
    return brand.strip(), ownership


def parse_category_rule(value: str) -> tuple[str, str]:
    if "=" not in value:
        raise argparse.ArgumentTypeError("Category rules must use Category=term one, term two.")
    category, terms = value.split("=", 1)
    return category.strip(), terms.strip()


def discovery_payload(discovery) -> dict:
    return {
        "products": len(discovery.products),
        "brands": discovery.brand_counts,
        "categories": discovery.category_counts,
        "unknown_brands": discovery.unknown_brand_counts,
        "uncategorized_products": [product.name for product in discovery.uncategorized_products],
        "link_summaries": discovery.link_summaries,
    }


def validate_workbook(path: Path) -> dict:
    workbook = load_workbook(path)
    sheet = workbook.active
    product_name_rows = [row for row in range(5, sheet.max_row + 1, 13)]
    product_count = 0
    hyperlink_count = 0
    categories: dict[str, int] = {}
    for row in product_name_rows:
        category = str(sheet.cell(row - 2, 1).value or "").strip()
        row_products = 0
        for col in range(1, 16, 2):
            cell = sheet.cell(row, col)
            if cell.value:
                product_count += 1
                row_products += 1
            if cell.hyperlink:
                hyperlink_count += 1
        if category and row_products:
            categories[category] = categories.get(category, 0) + row_products

    return {
        "path": str(path),
        "rows": sheet.max_row,
        "products": product_count,
        "product_hyperlinks": hyperlink_count,
        "embedded_images": len(getattr(sheet, "_images", [])),
        "categories": categories,
    }


def run(args: argparse.Namespace) -> dict:
    links = normalize_links(args.url)
    if not links:
        raise SystemExit("At least one URL is required.")

    first_discovery = discover_links(links)
    for brand, ownership in args.brand_rule:
        add_brand_rule(brand, ownership)
    for category, terms in args.category_rule:
        add_category_rule(category, terms)
    final_discovery = discover_links(links)

    if final_discovery.unknown_brand_counts:
        raise SystemExit(f"Unknown brands remain: {final_discovery.unknown_brand_counts}")
    if final_discovery.uncategorized_products:
        names = [product.name for product in final_discovery.uncategorized_products]
        raise SystemExit(f"Uncategorized products remain: {names}")

    build_request = BuildRequest(
        links=links,
        output_path=args.output,
        append_to=args.append_to,
        feature_alignment=args.feature_alignment,
    )
    workbook_path = run_build(build_request)
    result = {
        "input": {
            "urls": links,
            "output": str(args.output),
            "append_to": str(args.append_to) if args.append_to else "",
            "feature_alignment": args.feature_alignment,
            "brand_rules": [list(rule) for rule in args.brand_rule],
            "category_rules": [list(rule) for rule in args.category_rule],
        },
        "before_rules": discovery_payload(first_discovery),
        "after_rules": discovery_payload(final_discovery),
        "workbook": validate_workbook(workbook_path),
    }
    args.summary.write_text(json.dumps(result, indent=2), encoding="utf-8")
    return result


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Programmatic backend end-to-end workflow validation.")
    parser.add_argument("--url", action="append", required=True, help="Listing URL to scrape. Repeat for multiple URLs.")
    parser.add_argument("--output", type=Path, required=True, help="Workbook path to create.")
    parser.add_argument("--append-to", type=Path, help="Existing workbook to append to.")
    parser.add_argument(
        "--summary",
        type=Path,
        default=Path("scraper_outputs/app_backend_e2e_summary.json"),
        help="JSON summary path.",
    )
    parser.add_argument(
        "--brand-rule",
        action="append",
        default=[],
        type=parse_brand_rule,
        help="Apply a brand ownership rule, for example Thrustmaster=competitor.",
    )
    parser.add_argument(
        "--category-rule",
        action="append",
        default=[],
        type=parse_category_rule,
        help="Apply a category rule, for example Racing wheels=racing wheel,t248.",
    )
    parser.add_argument("--feature-alignment", choices=["llm", "raw"], default="llm")
    return parser.parse_args()


if __name__ == "__main__":
    print(json.dumps(run(parse_args()), indent=2))

from __future__ import annotations

import argparse
import json
import re
import shutil
from collections import OrderedDict
from copy import copy
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image as PILImage

from feature_alignment_llm import ProductFeatureInput, align_features_with_local_llm
from scraper_tools import (
    DEFAULT_RULES_PATH,
    CategoryProduct,
    ProductDetail,
    infer_brand,
    infer_product_category,
    load_parser_rules,
    make_session,
    read_links,
    scrape_category_pages,
    scrape_product_detail,
)


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_LINKS_PATH = ROOT / "links.txt"
DEFAULT_TEMPLATE_PATH = ROOT / "Competitors Analysis Blank Template.xlsx"
DEFAULT_OUTPUT_PATH = ROOT / "Competitors Analysis - Incredible Gaming.xlsx"
DEFAULT_OUTPUT_DIR = ROOT / "scraper_outputs"
OWN_BRAND_FILL = "FFF9D1A5"
REPORT_TITLE = "Incredible Connection Gaming Accessory Analysis"
PRODUCT_COLUMNS = [1, 3, 5, 7, 9, 11, 13, 15]
BLOCK_SOURCE_START = 3
BLOCK_SOURCE_END = 14
BLOCK_HEIGHT = 13
FEATURE_ROW_COUNT = 5
IMAGE_BOX_PIXELS = 260
PRODUCT_MAIN_COLUMN_WIDTH = 24
PRODUCT_SECONDARY_COLUMN_WIDTH = 23
MAX_LLM_TEXT_CHARS = 1200
MAX_LLM_SPEC_LINES = 20
PRODUCT_NAME_ROW_HEIGHT = 30.75
SPACER_ROW_HEIGHT = 28.5
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")


def safe_filename(value: str) -> str:
    value = re.sub(r"[^A-Za-z0-9_-]+", "_", value).strip("_")
    return value[:80] or "image"


def copy_block(source_ws, target_ws, target_start_row: int) -> None:
    row_offset = target_start_row - BLOCK_SOURCE_START
    for row in range(BLOCK_SOURCE_START, BLOCK_SOURCE_END + 1):
        target_row = row + row_offset
        target_ws.row_dimensions[target_row].height = source_ws.row_dimensions[row].height
        for col in range(1, 17):
            source_cell = source_ws.cell(row, col)
            target_cell = target_ws.cell(target_row, col)
            if isinstance(target_cell, MergedCell):
                continue
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            target_cell.number_format = source_cell.number_format
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.protection = copy(source_cell.protection)
            target_cell.value = source_cell.value
    for merged in source_ws.merged_cells.ranges:
        if merged.min_row < BLOCK_SOURCE_START or merged.max_row > BLOCK_SOURCE_END or merged.max_col > 16:
            continue
        shifted = (
            f"{get_column_letter(merged.min_col)}{merged.min_row + row_offset}:"
            f"{get_column_letter(merged.max_col)}{merged.max_row + row_offset}"
        )
        target_ws.merge_cells(shifted)
    copy_spacer_row(source_ws, target_ws, target_start_row + BLOCK_HEIGHT - 1)
    ensure_product_pair_merges(target_ws, target_start_row)
    style_generated_feature_rows(source_ws, target_ws, target_start_row)
    apply_feature_borders(target_ws, target_start_row)
    apply_block_alignment(target_ws, target_start_row)


def copy_spacer_row(source_ws, target_ws, target_row: int) -> None:
    source_row = BLOCK_SOURCE_END + 1 if is_blank_row(source_ws, BLOCK_SOURCE_END + 1) else BLOCK_SOURCE_END
    target_ws.row_dimensions[target_row].height = source_ws.row_dimensions[source_row].height or SPACER_ROW_HEIGHT
    for col in range(1, 17):
        source_cell = source_ws.cell(source_row, col)
        target_cell = target_ws.cell(target_row, col)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        target_cell.number_format = source_cell.number_format
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.protection = copy(source_cell.protection)
        target_cell.value = None
        target_cell.hyperlink = None


def is_blank_row(ws, row: int) -> bool:
    if row > ws.max_row:
        return False
    return all(ws.cell(row, col).value in (None, "") for col in range(1, 17))


def ensure_product_pair_merges(ws, start_row: int) -> None:
    for row in range(start_row + 7, start_row + 7 + FEATURE_ROW_COUNT):
        for col in PRODUCT_COLUMNS:
            cell_range = f"{get_column_letter(col)}{row}:{get_column_letter(col + 1)}{row}"
            if cell_range not in {str(merged) for merged in ws.merged_cells.ranges}:
                ws.merge_cells(cell_range)


def style_generated_feature_rows(source_ws, target_ws, start_row: int) -> None:
    feature_style_row = 10
    for row in range(start_row + 7, start_row + 7 + FEATURE_ROW_COUNT):
        target_ws.row_dimensions[row].height = source_ws.row_dimensions[feature_style_row].height
        for col in range(1, 17):
            source_cell = source_ws.cell(feature_style_row, col)
            target_cell = target_ws.cell(row, col)
            if isinstance(target_cell, MergedCell):
                continue
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            target_cell.number_format = source_cell.number_format
            target_cell.alignment = Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                text_rotation=source_cell.alignment.text_rotation,
                wrap_text=True,
                shrink_to_fit=source_cell.alignment.shrink_to_fit,
                indent=source_cell.alignment.indent,
            )
            target_cell.protection = copy(source_cell.protection)


def apply_feature_borders(ws, start_row: int) -> None:
    black_side = Side(style="thin", color="000000")
    full_border = Border(left=black_side, right=black_side, top=black_side, bottom=black_side)
    for row in range(start_row + 7, start_row + 7 + FEATURE_ROW_COUNT):
        for col in range(1, 17):
            ws.cell(row, col).border = full_border


def apply_block_alignment(ws, start_row: int) -> None:
    ws.row_dimensions[start_row + 2].height = PRODUCT_NAME_ROW_HEIGHT
    for col in range(1, 17):
        ws.cell(start_row + 2, col).alignment = copy(CENTER_ALIGNMENT)
        ws.cell(start_row + 6, col).alignment = copy(CENTER_ALIGNMENT)


def column_width_to_pixels(width: float | None) -> int:
    if not width:
        return 64
    return int(width * 7 + 5)


def points_to_pixels(points: float | None) -> int:
    if not points:
        return 20
    return int(points * 96 / 72)


def centered_image_anchor(ws, cell, width_px: int, height_px: int) -> OneCellAnchor:
    merged_width_px = sum(
        column_width_to_pixels(ws.column_dimensions[get_column_letter(col)].width)
        for col in (cell.column, cell.column + 1)
    )
    row_height_px = points_to_pixels(ws.row_dimensions[cell.row].height)
    left_offset_px = max(0, (merged_width_px - width_px) // 2)
    top_offset_px = max(0, (row_height_px - height_px) // 2)
    marker = AnchorMarker(
        col=cell.column - 1,
        colOff=pixels_to_EMU(left_offset_px),
        row=cell.row - 1,
        rowOff=pixels_to_EMU(top_offset_px),
    )
    ext = XDRPositiveSize2D(cx=pixels_to_EMU(width_px), cy=pixels_to_EMU(height_px))
    return OneCellAnchor(_from=marker, ext=ext)


def clear_product_slots(ws, start_row: int) -> None:
    for row in range(start_row + 1, start_row + 7 + FEATURE_ROW_COUNT):
        for col in range(1, 17):
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
            cell.hyperlink = None


def reset_to_header(workbook_path: Path):
    wb = load_workbook(workbook_path)
    ws = wb.active
    template_ws = load_workbook(workbook_path).active
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row >= 3:
            ws.unmerge_cells(str(merged))
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)
    for col in range(1, 17):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = (
            PRODUCT_MAIN_COLUMN_WIDTH if col % 2 else PRODUCT_SECONDARY_COLUMN_WIDTH
        )
    return wb, ws


def find_next_block_start(ws) -> int:
    populated_rows = [
        cell.row
        for row in ws.iter_rows(min_row=BLOCK_SOURCE_START, max_col=16)
        for cell in row
        if cell.value not in (None, "")
    ]
    if not populated_rows:
        return BLOCK_SOURCE_START
    highest_row = max(populated_rows)
    block_count = ((highest_row - BLOCK_SOURCE_START) // BLOCK_HEIGHT) + 1
    return BLOCK_SOURCE_START + (block_count * BLOCK_HEIGHT)


def open_workbook_target(args):
    template_wb = load_workbook(args.template)
    template_ws = template_wb.active
    if args.append_to:
        source_path = args.append_to
        if args.output.resolve() != source_path.resolve():
            args.output.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source_path, args.output)
            source_path = args.output
        wb = load_workbook(source_path)
        ws = wb.active
        for col in range(1, 17):
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = (
                PRODUCT_MAIN_COLUMN_WIDTH if col % 2 else PRODUCT_SECONDARY_COLUMN_WIDTH
            )
        return template_ws, wb, ws, find_next_block_start(ws), "append"

    wb, ws = reset_to_header(args.template)
    ws["A1"] = REPORT_TITLE
    ws["A2"] = "Buyer / Account"
    ws["O1"] = args.month
    return template_ws, wb, ws, BLOCK_SOURCE_START, "fresh"


def get_product_detail(product: CategoryProduct, session: requests.Session) -> ProductDetail:
    try:
        return scrape_product_detail(product.product_url, session=session)
    except Exception:
        return ProductDetail(
            product_url=product.product_url,
            title=product.name,
            image_url=product.image_url,
            overview="",
            description="",
            specifications=[],
        )


def product_feature_input(
    product_id: str,
    product: CategoryProduct,
    detail: ProductDetail,
    category: str,
    brand: str,
) -> ProductFeatureInput:
    return ProductFeatureInput(
        product_id=product_id,
        name=detail.title or product.name,
        brand=brand,
        category=category,
        overview=truncate_text(detail.overview),
        description=truncate_text(detail.description),
        specifications=[truncate_text(spec) for spec in detail.specifications[:MAX_LLM_SPEC_LINES]],
    )


def truncate_text(value: str, limit: int = MAX_LLM_TEXT_CHARS) -> str:
    value = value.strip()
    if len(value) <= limit:
        return value
    return value[:limit].rsplit(" ", 1)[0].strip()


def download_image(image_url: str, image_dir: Path, name: str, session: requests.Session) -> Path | None:
    if not image_url:
        return None
    image_dir.mkdir(parents=True, exist_ok=True)
    try:
        response = session.get(image_url, timeout=45)
        response.raise_for_status()
        raw_path = image_dir / f"{safe_filename(name)}.img"
        raw_path.write_bytes(response.content)
        with PILImage.open(raw_path) as image:
            image = image.convert("RGBA")
            image.thumbnail((IMAGE_BOX_PIXELS, IMAGE_BOX_PIXELS))
            canvas = PILImage.new("RGBA", (IMAGE_BOX_PIXELS, IMAGE_BOX_PIXELS), (255, 255, 255, 0))
            left = (IMAGE_BOX_PIXELS - image.width) // 2
            top = (IMAGE_BOX_PIXELS - image.height) // 2
            canvas.alpha_composite(image, (left, top))
            output_path = image_dir / f"{safe_filename(name)}.png"
            canvas.save(output_path)
        raw_path.unlink(missing_ok=True)
        return output_path
    except Exception:
        return None


def apply_own_brand_fill(ws, start_row: int, col: int) -> None:
    fill = PatternFill("solid", fgColor=OWN_BRAND_FILL)
    for row in range(start_row + 2, start_row + 7 + FEATURE_ROW_COUNT):
        for current_col in (col, col + 1):
            ws.cell(row, current_col).fill = copy(fill)


def display_sku(raw_sku: str) -> str:
    if not raw_sku:
        return ""
    stripped = raw_sku.lstrip("0")
    return stripped or "0"


def price_number(value: str) -> float:
    if not value:
        return float("inf")
    cleaned = re.sub(r"[^0-9.,]", "", value).replace(" ", "")
    if "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return float("inf")


def everyday_price_number(product: CategoryProduct) -> float:
    return price_number(product.old_price or product.final_price)


def fill_product_slot(
    ws,
    start_row: int,
    col: int,
    product: CategoryProduct,
    category: str,
    brand: str,
    own_brands: set[str],
    detail: ProductDetail,
    aligned_features: list[str],
    session: requests.Session,
    image_dir: Path,
) -> None:
    image_url = detail.image_url or product.image_url
    ws.cell(start_row, 1).value = category

    image_path = download_image(image_url or product.image_url, image_dir, product.name, session)
    picture_cell = ws.cell(start_row + 1, col)
    picture_cell.value = ""
    if image_path:
        image = ExcelImage(str(image_path))
        image.width = IMAGE_BOX_PIXELS
        image.height = IMAGE_BOX_PIXELS
        image.anchor = centered_image_anchor(ws, picture_cell, IMAGE_BOX_PIXELS, IMAGE_BOX_PIXELS)
        ws.add_image(image)
    elif image_url:
        picture_cell.value = "Image URL"
        picture_cell.hyperlink = image_url
        picture_cell.style = "Hyperlink"

    name_cell = ws.cell(start_row + 2, col)
    name_cell.value = product.name
    name_cell.hyperlink = product.product_url
    name_cell.style = "Hyperlink"
    name_cell.alignment = copy(CENTER_ALIGNMENT)

    ws.cell(start_row + 3, col).value = display_sku(product.sku)
    ws.cell(start_row + 4, col).value = "Everyday pricing"
    ws.cell(start_row + 4, col + 1).value = "Product on promo?"
    ws.cell(start_row + 5, col + 1).value = product.promo
    ws.cell(start_row + 6, col).value = product.old_price or product.final_price
    ws.cell(start_row + 6, col + 1).value = product.final_price if product.old_price else "N/A"
    for offset, feature in enumerate(aligned_features[:FEATURE_ROW_COUNT], start=7):
        ws.cell(start_row + offset, col).value = feature

    if brand in own_brands:
        apply_own_brand_fill(ws, start_row, col)


def aligned_feature_lines(alignment, product_ids: list[str]) -> dict[str, list[str]]:
    lines = {product_id: [""] * FEATURE_ROW_COUNT for product_id in product_ids}
    for row_index, row in enumerate(alignment.rows[:FEATURE_ROW_COUNT]):
        for product_id in product_ids:
            value = clean_feature_value(row.values.get(product_id, ""))
            lines[product_id][row_index] = value
    return lines


def clean_feature_value(value: str) -> str:
    value = str(value or "").strip()
    return re.sub(r"^feature\s*\d+\s*:\s*", "", value, flags=re.IGNORECASE).strip()


def align_chunk_features(
    chunk: list[CategoryProduct],
    category: str,
    details: dict[str, ProductDetail],
    parser_rules,
    llm_config: Path,
    mode: str,
) -> tuple[dict[str, list[str]], dict]:
    product_ids = [f"p{index + 1}" for index in range(len(chunk))]
    products = []
    for product_id, product in zip(product_ids, chunk):
        brand = infer_brand(product.name, parser_rules)
        detail = details[product.product_url]
        products.append(product_feature_input(product_id, product, detail, category, brand))

    if mode == "raw":
        rows = {product_id: [""] * FEATURE_ROW_COUNT for product_id in product_ids}
        for product_id, feature_input in zip(product_ids, products):
            source = feature_input.specifications
            if not source and feature_input.overview:
                source = [part.strip(" -") for part in re.split(r"(?<=[.;])\s+", feature_input.overview) if part.strip(" -")]
            if not source and feature_input.description:
                source = [part.strip(" -") for part in re.split(r"(?<=[.;])\s+", feature_input.description) if part.strip(" -")]
            rows[product_id] = source[:FEATURE_ROW_COUNT] + [""] * max(0, FEATURE_ROW_COUNT - len(source))
        return rows, {"category": category, "source": "raw", "warnings": []}

    alignment = align_features_with_local_llm(products, config_path=llm_config, allow_fallback=True)
    return aligned_feature_lines(alignment, product_ids), {
        "category": category,
        "source": alignment.source,
        "warnings": alignment.warnings,
        "products": [product.name for product in chunk],
    }


def collect_products(links: list[str], parser_rules) -> tuple[list[CategoryProduct], list[dict]]:
    session = make_session()
    all_products: list[CategoryProduct] = []
    summaries: list[dict] = []
    seen: set[str] = set()
    for link in links:
        result = scrape_category_pages(link, parser_rules=parser_rules, session=session)
        summaries.append(
            {
                "source_url": link,
                "validation": result.validation,
                "brand_counts": result.brand_counts,
                "category_counts": result.category_counts,
                "product_count": len(result.products),
            }
        )
        for product in result.products:
            key = product.product_url or product.name.casefold()
            if key in seen:
                continue
            seen.add(key)
            all_products.append(product)
    return all_products, summaries


def group_products(products: list[CategoryProduct], parser_rules) -> OrderedDict[str, list[CategoryProduct]]:
    groups: OrderedDict[str, list[CategoryProduct]] = OrderedDict()
    for product in products:
        category = infer_product_category(product.name, parser_rules)
        groups.setdefault(category, []).append(product)
    for category, items in groups.items():
        groups[category] = sorted(items, key=lambda product: (everyday_price_number(product), product.name.casefold()))
    return groups


def build_workbook(args) -> None:
    parser_rules = load_parser_rules(args.rules)
    links = [args.url] if args.url else read_links(args.links)
    products, summaries = collect_products(links, parser_rules)
    grouped = group_products(products, parser_rules)

    template_ws, wb, ws, current_row, workbook_mode = open_workbook_target(args)
    session = make_session()
    image_dir = args.output_dir / "embedded_images"
    own_brands = set(parser_rules.own_brands)
    details: dict[str, ProductDetail] = {}
    feature_alignment_runs: list[dict] = []

    for category, category_products in grouped.items():
        for chunk_start in range(0, len(category_products), len(PRODUCT_COLUMNS)):
            copy_block(template_ws, ws, current_row)
            clear_product_slots(ws, current_row)
            chunk = category_products[chunk_start : chunk_start + len(PRODUCT_COLUMNS)]
            for product in chunk:
                details[product.product_url] = get_product_detail(product, session)
            ws.cell(current_row, 1).value = category
            if chunk_start:
                ws.cell(current_row, 1).value = f"{category} (continued)"
            display_category = ws.cell(current_row, 1).value
            feature_lines, alignment_summary = align_chunk_features(
                chunk,
                category,
                details,
                parser_rules,
                args.llm_config,
                args.feature_alignment,
            )
            alignment_summary["display_category"] = display_category
            alignment_summary["row"] = current_row
            feature_alignment_runs.append(alignment_summary)
            for product_index, (col, product) in enumerate(zip(PRODUCT_COLUMNS, chunk), start=1):
                brand = infer_brand(product.name, parser_rules)
                product_id = f"p{product_index}"
                fill_product_slot(
                    ws,
                    current_row,
                    col,
                    product,
                    display_category,
                    brand,
                    own_brands,
                    details[product.product_url],
                    feature_lines[product_id],
                    session,
                    image_dir,
                )
            current_row += BLOCK_HEIGHT

    if current_row <= ws.max_row:
        ws.delete_rows(current_row, ws.max_row - current_row + 1)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)

    summary_path = args.output_dir / "combined_competitor_analysis_summary.json"
    args.output_dir.mkdir(exist_ok=True)
    summary_path.write_text(
        json.dumps(
            {
                "output": str(args.output),
                "workbook_mode": workbook_mode,
                "append_to": str(args.append_to) if args.append_to else "",
                "own_brands": sorted(own_brands),
                "input_links": links,
                "link_summaries": summaries,
                "feature_alignment": {
                    "mode": args.feature_alignment,
                    "llm_config": str(args.llm_config),
                    "runs": feature_alignment_runs,
                },
                "combined_categories": {category: len(items) for category, items in grouped.items()},
                "combined_brands": dict(sorted({brand: sum(1 for product in products if infer_brand(product.name, parser_rules) == brand) for brand in {infer_brand(product.name, parser_rules) for product in products}}.items())),
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    print(f"Wrote {args.output}")
    print(f"Wrote {summary_path}")


def parse_args():
    parser = argparse.ArgumentParser(description="Build a combined competitor analysis Excel workbook from category links.")
    parser.add_argument("--url", help="Optional single URL. Defaults to all URLs in links.txt.")
    parser.add_argument("--links", type=Path, default=DEFAULT_LINKS_PATH)
    parser.add_argument("--rules", type=Path, default=DEFAULT_RULES_PATH)
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE_PATH)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_PATH)
    parser.add_argument("--append-to", type=Path, help="Existing competitor analysis workbook to append to instead of starting fresh.")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--llm-config", type=Path, default=ROOT / "config" / "local_llm.json")
    parser.add_argument(
        "--feature-alignment",
        choices=["llm", "raw"],
        default="llm",
        help="Use the local LLM to align comparable feature rows, or write raw scraped feature order.",
    )
    parser.add_argument("--month", default=datetime.now().strftime("%b-%y"), help="Report month shown in the workbook header.")
    return parser.parse_args()


if __name__ == "__main__":
    build_workbook(parse_args())

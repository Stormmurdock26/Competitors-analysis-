from __future__ import annotations

import asyncio
import json
import os
import re
import time
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Callable

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
LINKS_PATH = ROOT / "links.txt"
TEMPLATE_PATH = ROOT / "Competitors Analysis Blank Template.xlsx"
OUTPUT_DIR = ROOT / "scraper_outputs"
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"


@dataclass
class Product:
    name: str
    sku: str
    final_price: str
    old_price: str
    promo: str
    product_url: str
    image_url: str
    source_product_id: str


@dataclass
class ScraperResult:
    scraper: str
    status: str
    elapsed_seconds: float
    product_count: int
    field_score: float
    notes: str
    products: list[Product]


def read_links() -> list[str]:
    return [line.strip() for line in LINKS_PATH.read_text(encoding="utf-8").splitlines() if line.strip()]


def fetch_html(url: str) -> str:
    response = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=45)
    response.raise_for_status()
    return response.text


def text_clean(value: str | None) -> str:
    if not value:
        return ""
    return " ".join(value.split())


def normalize_price(value: str | None) -> str:
    value = text_clean(value)
    if not value:
        return ""
    try:
        amount = float(value)
    except ValueError:
        return value
    if amount.is_integer():
        return f"R{int(amount):,}".replace(",", " ")
    return f"R{amount:,.2f}".replace(",", " ")


def dedupe(products: list[Product]) -> list[Product]:
    seen: set[tuple[str, str]] = set()
    unique: list[Product] = []
    for product in products:
        key = (product.name.casefold(), product.product_url)
        if not product.name or key in seen:
            continue
        seen.add(key)
        unique.append(product)
    return unique


def score_products(products: list[Product]) -> float:
    fields = ("name", "sku", "final_price", "old_price", "promo", "product_url", "image_url")
    if not products:
        return 0.0
    filled = 0
    total = len(products) * len(fields)
    for product in products:
        for field in fields:
            if getattr(product, field):
                filled += 1
    return round(filled / total, 3)


def extract_with_bs4(html: str) -> list[Product]:
    soup = BeautifulSoup(html, "lxml")
    products: list[Product] = []
    for item in soup.select("li.item.product.product-item"):
        link = item.select_one("a.product-item-link")
        if not link:
            continue
        final = item.select_one('[data-price-type="finalPrice"]')
        old = item.select_one('[data-price-type="oldPrice"]')
        form = item.select_one("form[data-product-sku]")
        image = item.select_one("img.product-image-photo")
        source_id = ""
        price_box = item.select_one("[data-product-id]")
        if price_box:
            source_id = price_box.get("data-product-id", "")
        old_price = normalize_price(old.get("data-price-amount") if old else "")
        final_price = normalize_price(final.get("data-price-amount") if final else "")
        products.append(
            Product(
                name=text_clean(link.get_text(" ", strip=True)),
                sku=(form.get("data-product-sku", "") if form else ""),
                final_price=final_price,
                old_price=old_price,
                promo="Yes" if old_price else "No",
                product_url=link.get("href", ""),
                image_url=image.get("src", "") if image else "",
                source_product_id=source_id,
            )
        )
    return dedupe(products)


def extract_with_lxml(html: str) -> list[Product]:
    from lxml import html as lxml_html

    doc = lxml_html.fromstring(html)
    products: list[Product] = []
    for item in doc.cssselect("li.item.product.product-item"):
        links = item.cssselect("a.product-item-link")
        if not links:
            continue
        link = links[0]
        final = item.cssselect('[data-price-type="finalPrice"]')
        old = item.cssselect('[data-price-type="oldPrice"]')
        form = item.cssselect("form[data-product-sku]")
        image = item.cssselect("img.product-image-photo")
        price_box = item.cssselect("[data-product-id]")
        old_price = normalize_price(old[0].get("data-price-amount") if old else "")
        final_price = normalize_price(final[0].get("data-price-amount") if final else "")
        products.append(
            Product(
                name=text_clean(link.text_content()),
                sku=form[0].get("data-product-sku", "") if form else "",
                final_price=final_price,
                old_price=old_price,
                promo="Yes" if old_price else "No",
                product_url=link.get("href", ""),
                image_url=image[0].get("src", "") if image else "",
                source_product_id=price_box[0].get("data-product-id", "") if price_box else "",
            )
        )
    return dedupe(products)


def extract_with_selectolax(html: str) -> list[Product]:
    from selectolax.parser import HTMLParser

    tree = HTMLParser(html)
    products: list[Product] = []
    for item in tree.css("li.item.product.product-item"):
        link = item.css_first("a.product-item-link")
        if not link:
            continue
        final = item.css_first('[data-price-type="finalPrice"]')
        old = item.css_first('[data-price-type="oldPrice"]')
        form = item.css_first("form[data-product-sku]")
        image = item.css_first("img.product-image-photo")
        price_box = item.css_first("[data-product-id]")
        old_price = normalize_price(old.attributes.get("data-price-amount", "") if old else "")
        final_price = normalize_price(final.attributes.get("data-price-amount", "") if final else "")
        products.append(
            Product(
                name=text_clean(link.text()),
                sku=form.attributes.get("data-product-sku", "") if form else "",
                final_price=final_price,
                old_price=old_price,
                promo="Yes" if old_price else "No",
                product_url=link.attributes.get("href", ""),
                image_url=image.attributes.get("src", "") if image else "",
                source_product_id=price_box.attributes.get("data-product-id", "") if price_box else "",
            )
        )
    return dedupe(products)


def extract_with_scrapy(html: str) -> list[Product]:
    from scrapy.selector import Selector

    selector = Selector(text=html)
    products: list[Product] = []
    for item in selector.css("li.item.product.product-item"):
        name = text_clean(" ".join(item.css("a.product-item-link::text").getall()))
        if not name:
            continue
        old_price = normalize_price(item.css('[data-price-type="oldPrice"]::attr(data-price-amount)').get(""))
        final_price = normalize_price(item.css('[data-price-type="finalPrice"]::attr(data-price-amount)').get(""))
        products.append(
            Product(
                name=name,
                sku=item.css("form[data-product-sku]::attr(data-product-sku)").get("") or "",
                final_price=final_price,
                old_price=old_price,
                promo="Yes" if old_price else "No",
                product_url=item.css("a.product-item-link::attr(href)").get("") or "",
                image_url=item.css("img.product-image-photo::attr(src)").get("") or "",
                source_product_id=item.css("[data-product-id]::attr(data-product-id)").get("") or "",
            )
        )
    return dedupe(products)


async def scrape_playwright(url: str) -> list[Product]:
    from playwright.async_api import async_playwright

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page = await browser.new_page(user_agent=USER_AGENT)
        await page.goto(url, wait_until="networkidle", timeout=60000)
        html = await page.content()
        await browser.close()
    return extract_with_bs4(html)


def scrape_firecrawl(url: str) -> tuple[list[Product], str]:
    api_key = os.environ.get("FIRECRAWL_API_KEY")
    if not api_key:
        return [], "Skipped: FIRECRAWL_API_KEY is not set."
    try:
        from firecrawl import FirecrawlApp
    except Exception as exc:
        return [], f"Skipped: firecrawl package import failed: {exc}"
    try:
        app = FirecrawlApp(api_key=api_key)
        result = app.scrape_url(url, formats=["html"])
        html = ""
        if isinstance(result, dict):
            html = result.get("html") or result.get("content") or ""
        else:
            html = getattr(result, "html", "") or getattr(result, "content", "")
        if not html:
            return [], "Firecrawl ran but returned no HTML content to parse."
        return extract_with_bs4(html), "Firecrawl HTML parsed with the shared product-card extractor."
    except Exception as exc:
        return [], f"Firecrawl failed: {exc}"


def run_timed(name: str, callback: Callable[[], list[Product]]) -> ScraperResult:
    start = time.perf_counter()
    try:
        products = callback()
        status = "ok"
        notes = ""
    except Exception as exc:
        products = []
        status = "failed"
        notes = str(exc)
    elapsed = round(time.perf_counter() - start, 3)
    return ScraperResult(name, status, elapsed, len(products), score_products(products), notes, products)


def product_slots() -> list[tuple[int, int]]:
    starts = [3, 15, 27, 39, 51, 63]
    cols = [1, 3, 5, 7, 9, 11, 13, 15]
    return [(start, col) for start in starts for col in cols]


def write_excel(result: ScraperResult, url: str) -> Path:
    OUTPUT_DIR.mkdir(exist_ok=True)
    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", result.scraper).strip("_")
    out_path = OUTPUT_DIR / f"filled_{safe_name}.xlsx"
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active
    slots = product_slots()
    category = "Incredible Connection - Gaming Accessories"
    for index, product in enumerate(result.products[: len(slots)]):
        start, col = slots[index]
        ws.cell(start, 1).value = category
        picture_cell = ws.cell(start + 1, col)
        picture_cell.value = "Image URL" if product.image_url else ""
        if product.image_url:
            picture_cell.hyperlink = product.image_url
            picture_cell.style = "Hyperlink"
        name_cell = ws.cell(start + 2, col)
        name_cell.value = product.name
        if product.product_url:
            name_cell.hyperlink = product.product_url
            name_cell.style = "Hyperlink"
        ws.cell(start + 3, col).value = product.sku
        ws.cell(start + 4, col).value = "Everyday pricing"
        ws.cell(start + 4, col + 1).value = "Product on promo?"
        ws.cell(start + 5, col + 1).value = product.promo
        ws.cell(start + 6, col).value = product.old_price or product.final_price
        ws.cell(start + 6, col + 1).value = product.final_price if product.old_price else "N/A"
        ws.cell(start + 7, col).value = f"Scraper: {result.scraper}"
        ws.cell(start + 8, col).value = f"Product ID: {product.source_product_id}" if product.source_product_id else ""
        ws.cell(start + 9, col).value = f"Source link: {url}"
    wb.save(out_path)
    return out_path


def write_summary(results: list[ScraperResult], outputs: dict[str, Path]) -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)
    summary = []
    summary.append("# Scraper Comparison Results")
    summary.append("")
    summary.append(f"Input link: `{read_links()[0]}`")
    summary.append("")
    summary.append("| Scraper | Status | Products | Field score | Seconds | Excel output | Notes |")
    summary.append("| --- | --- | ---: | ---: | ---: | --- | --- |")
    for result in results:
        output = outputs.get(result.scraper)
        output_name = output.name if output else ""
        notes = result.notes.replace("|", "\\|")
        summary.append(
            f"| {result.scraper} | {result.status} | {result.product_count} | "
            f"{result.field_score:.3f} | {result.elapsed_seconds:.3f} | {output_name} | {notes} |"
        )
    summary.append("")
    summary.append("Field score measures how many of the normalized fields were populated: name, SKU, final price, old price, promo flag, product URL, and image URL.")
    (OUTPUT_DIR / "scraper_comparison_summary.md").write_text("\n".join(summary) + "\n", encoding="utf-8")
    json_data = []
    for result in results:
        row = asdict(result)
        row["products"] = [asdict(product) for product in result.products]
        json_data.append(row)
    (OUTPUT_DIR / "scraper_comparison_results.json").write_text(json.dumps(json_data, indent=2), encoding="utf-8")


async def main() -> None:
    links = read_links()
    if not links:
        raise SystemExit("No links found in links.txt")
    url = links[0]
    html = fetch_html(url)
    results: list[ScraperResult] = []
    results.append(run_timed("beautifulsoup", lambda: extract_with_bs4(html)))
    results.append(run_timed("lxml", lambda: extract_with_lxml(html)))
    results.append(run_timed("selectolax", lambda: extract_with_selectolax(html)))
    results.append(run_timed("scrapy_selector", lambda: extract_with_scrapy(html)))

    start = time.perf_counter()
    try:
        playwright_products = await scrape_playwright(url)
        results.append(
            ScraperResult(
                "playwright",
                "ok",
                round(time.perf_counter() - start, 3),
                len(playwright_products),
                score_products(playwright_products),
                "",
                playwright_products,
            )
        )
    except Exception as exc:
        results.append(ScraperResult("playwright", "failed", round(time.perf_counter() - start, 3), 0, 0.0, str(exc), []))

    start = time.perf_counter()
    firecrawl_products, firecrawl_notes = scrape_firecrawl(url)
    results.append(
        ScraperResult(
            "firecrawl",
            "ok" if firecrawl_products else "skipped",
            round(time.perf_counter() - start, 3),
            len(firecrawl_products),
            score_products(firecrawl_products),
            firecrawl_notes,
            firecrawl_products,
        )
    )

    outputs: dict[str, Path] = {}
    for result in results:
        if result.products:
            outputs[result.scraper] = write_excel(result, url)
    write_summary(results, outputs)
    print(json.dumps({result.scraper: asdict(result) | {"products": result.product_count} for result in results}, indent=2))
    print(f"Outputs written to {OUTPUT_DIR}")


if __name__ == "__main__":
    asyncio.run(main())

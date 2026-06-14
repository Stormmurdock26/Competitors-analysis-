from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from build_competitor_analysis import (
    BLOCK_HEIGHT,
    BLOCK_SOURCE_END,
    BLOCK_SOURCE_START,
    CENTER_ALIGNMENT,
    DEFAULT_TEMPLATE_PATH,
    FEATURE_ROW_COUNT,
    PRODUCT_COLUMNS,
    PRODUCT_MAIN_COLUMN_WIDTH,
    PRODUCT_NAME_ROW_HEIGHT,
    PRODUCT_SECONDARY_COLUMN_WIDTH,
    OWN_BRAND_FILL,
    SPACER_ROW_HEIGHT,
    apply_block_alignment,
    apply_feature_borders,
    copy_block,
    ensure_product_pair_merges,
    style_generated_feature_rows,
)


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_BLOCKS = 6
BLANK_TEMPLATE_TITLE = "Competitor Analysis"
NO_FILL = PatternFill(fill_type=None)


def reset_body(ws) -> None:
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row >= BLOCK_SOURCE_START:
            ws.unmerge_cells(str(merged))
    if ws.max_row >= BLOCK_SOURCE_START:
        ws.delete_rows(BLOCK_SOURCE_START, ws.max_row - BLOCK_SOURCE_START + 1)
    ws._images = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.hyperlink = None


def set_column_widths(ws) -> None:
    for col in range(1, 17):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = (
            PRODUCT_MAIN_COLUMN_WIDTH if col % 2 else PRODUCT_SECONDARY_COLUMN_WIDTH
        )


def clear_runtime_product_formatting(ws, start_row: int) -> None:
    for row in range(start_row + 1, start_row + 7 + FEATURE_ROW_COUNT):
        for col in range(1, 17):
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                continue
            cell.hyperlink = None
            if is_owned_brand_fill(cell):
                cell.fill = copy(NO_FILL)


def is_owned_brand_fill(cell) -> bool:
    fg_color = cell.fill.fgColor
    return cell.fill.fill_type == "solid" and fg_color.type == "rgb" and fg_color.rgb == OWN_BRAND_FILL


def write_blank_block(ws, start_row: int) -> None:
    ws.cell(start_row, 1).value = "Category"
    for col in PRODUCT_COLUMNS:
        ws.cell(start_row + 1, col).value = "Picture"
        ws.cell(start_row + 1, col).alignment = copy(CENTER_ALIGNMENT)
        ws.cell(start_row + 2, col).value = "Product name"
        ws.cell(start_row + 3, col).value = "Article / SKU"
        ws.cell(start_row + 4, col).value = "Everyday pricing"
        ws.cell(start_row + 4, col + 1).value = "Product on promo?"
        ws.cell(start_row + 5, col + 1).value = "Yes / No"
        ws.cell(start_row + 6, col).value = "Everyday price"
        ws.cell(start_row + 6, col + 1).value = "Promo price / N/A"
        for index in range(FEATURE_ROW_COUNT):
            ws.cell(start_row + 7 + index, col).value = f"Feature / spec {index + 1}"

    for row in range(start_row + 1, start_row + 7 + FEATURE_ROW_COUNT):
        for col in range(1, 17):
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                continue
            cell.hyperlink = None

    ws.row_dimensions[start_row + 2].height = PRODUCT_NAME_ROW_HEIGHT
    ws.row_dimensions[start_row + BLOCK_HEIGHT - 1].height = SPACER_ROW_HEIGHT


def build_template(source_path: Path, output_path: Path, blocks: int) -> None:
    source_wb = load_workbook(source_path)
    source_ws = source_wb.active
    wb = load_workbook(source_path)
    ws = wb.active

    reset_body(ws)
    set_column_widths(ws)
    ws["A1"] = BLANK_TEMPLATE_TITLE
    ws["A2"] = "Buyer / Account"
    ws["O1"] = "Month"

    current_row = BLOCK_SOURCE_START
    for _ in range(blocks):
        copy_block(source_ws, ws, current_row)
        ensure_product_pair_merges(ws, current_row)
        style_generated_feature_rows(source_ws, ws, current_row)
        apply_feature_borders(ws, current_row)
        apply_block_alignment(ws, current_row)
        clear_runtime_product_formatting(ws, current_row)
        write_blank_block(ws, current_row)
        current_row += BLOCK_HEIGHT

    for row in range(BLOCK_SOURCE_START, current_row):
        for col in range(1, 17):
            cell = ws.cell(row, col)
            if not isinstance(cell, MergedCell):
                cell.hyperlink = None

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Wrote {output_path}")


def parse_args():
    parser = argparse.ArgumentParser(description="Rebuild the blank competitor analysis template with current formatting.")
    parser.add_argument("--source", type=Path, default=DEFAULT_TEMPLATE_PATH)
    parser.add_argument("--output", type=Path, default=DEFAULT_TEMPLATE_PATH)
    parser.add_argument("--blocks", type=int, default=DEFAULT_BLOCKS)
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    build_template(args.source, args.output, args.blocks)
